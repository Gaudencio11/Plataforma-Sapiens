from django.shortcuts import render, get_object_or_404, redirect
from .models import Event, Room, Program,  AddIcon, SelectIcon, User

from django.core.mail import send_mail


#Excel sheet
from openpyxl import Workbook, workbook
from django.http import HttpResponse
from datetime import datetime
from datetime import timedelta

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#not in use at the moment

"""if AutorizedCpfs.objects.all() == True:
    lista_bd = AutorizedCpfs.objects.all()
    lista_cpf = []
    for item in lista_bd:
        lista_cpf.append(item.cpf)"""


def menuView(request): 
    if not request.user.is_authenticated:
        return redirect('/') 


    """send_mail('Primeiro django test mail', 'Se você está recebendo esse email é porquê você conseguiu', 'marcos.teste083@gmail.com', ['marcos.otavio10@gmail.com',])"""



    """if AutorizedCpfs.objects.all() == True:
        lista_bd = AutorizedCpfs.objects.all()
        lista_cpf = []
        for item in lista_bd:
            lista_cpf.append(item.cpf)

        if request.user.last_name not in lista_cpf: 

            
            if request.user.is_superuser:
                pass
            else:
                user_data = [request.user.username, request.user.last_name] 
                request.user.delete()
                return redirect('/cpf-invalido/')"""
        

    event = Event.objects.all()[0]

    

    event_name = event.slug
    
    #dar uma olhada na linha abaixo depois (apenas rooms da sapiens)
    rooms = Room.objects.all()
    
    return render(request, 'hall.html', {'event':event, 'rooms':rooms, 'slug':event_name})


    #slug2 é o slug da sala
def daysView(request, slug1, slug2):

    if not request.user.is_authenticated:
        return redirect('/')

    event = Event.objects.all()[0]
    

    #sala específica que a pessoa entrou
    sala = get_object_or_404(Room, slug=slug2) 

    
    rooms = Room.objects.all()
    
    sala_days= [x for x in range(1, event.days+1)]

    if sala.days == 1:
        return redirect(roomsGenericView, slug1, slug2, 'unico')

    return render(request, 'days.html', { "sala_days":sala_days, "room":sala, 'event':event, 'rooms':rooms})

#slug3 é o dia
def roomsGenericView(request, slug1, slug2, slug3):

    if not request.user.is_authenticated:
        return redirect('/')


    event = Event.objects.all()[0]
    

    sala = get_object_or_404(Room, slug=slug2)

    #Apenas programas da sala específica
    #Se não for dia único ele pega os programas específicos por dia
    #Se for dia único, ele pega apenas o programa daquela sala
    if slug3 == 'unico':
        room_programs = Program.objects.filter(room=sala.id)
    else:
        room_programs = Program.objects.filter(room=sala.id, day =slug3)
    


    #dar uma olhada na linha abaixo depois (apenas rooms da sapiens)
    rooms = Room.objects.all()

    return render(request, 'rooms.html', { 'event':event, 'programs': room_programs, 'room':sala, 'rooms':rooms})

#slug1 is the event
def padletView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    rooms = Room.objects.all()
    event = Event.objects.all()[0] 
    


    return render(request, 'padlet.html', {'event':event, 'rooms':rooms, 'slug1':slug1})


def desopilarView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    event = Event.objects.all()[0]
    
    rooms = Room.objects.all()  

    return render(request, 'desopilar.html', {'event': event, 'rooms':rooms, 'slug1':slug1})

def programacaoView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    event = Event.objects.all()[0]
    
    rooms = Room.objects.all()  

    return render(request, 'programacao.html', {'event': event, 'rooms':rooms, 'slug1':slug1})


def apoiadoresView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    event = Event.objects.all()[0]
    
    rooms = Room.objects.all() 


    return render(request, 'apoiadores.html', {'event': event, 'rooms':rooms, 'slug1':slug1})





def get_excell_users(request):
    
    

    users_list =  User.objects.all()

    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename={event}-usuarios.xlsx'.format(
        event = Event.objects.all()[0],
    )
    
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Usuários"

    columns = ['Posição','Nome completo','Curso', 'Email']
    row_num = 1

    #This atribue value to the columns headers
    for col_number, column_title in enumerate(columns,1):
        cell = worksheet.cell(row = row_num, column=col_number)
        cell.value = column_title

    #Define the data for each cell in the row
    for user in users_list:
        row_num +=1

        row=[
            row_num-1,
            str(user.username),
            user.last_name,
            user.email,
        ]

        for col_number, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_number)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

    #for fix the width of th column
    dimensions = [7,55, 20, 30]
    for col_num, width in enumerate(dimensions, 1):
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = width

    workbook.save(response)

    return response















"""def testeStatic1(request):

    return render(request, 'teste1.html', {})

def testeStatic2(request):

    return render(request, 'teste2.html', {})"""



